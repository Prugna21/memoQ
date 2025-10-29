import pandas as pd
import openpyxl
import sys
import os

def find_yellow_cells(file_name):
    """Find cells that are highlighted in yellow"""
    print("Looking for yellow cells...")
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_name, data_only=False)
    sheet = workbook.active
    
    yellow_cells = {}
    
    # Go through each row and column
    row_number = 1
    for row in sheet.iter_rows():
        col_number = 1
        for cell in row:
            # Check if cell is yellow
            if cell.fill and cell.fill.start_color:
                color = str(cell.fill.start_color.index)
                if 'FFFF' in color.upper():  # Yellow color
                    if row_number not in yellow_cells:
                        yellow_cells[row_number] = {}
                    yellow_cells[row_number][col_number] = cell.value
            col_number = col_number + 1
        row_number = row_number + 1
    
    workbook.close()
    print(f"Found yellow cells in {len(yellow_cells)} rows")
    return yellow_cells

def create_memoq_file(input_file, output_file):
    """Create a file for memoQ with only yellow highlighted rows"""
    
    # Step 1: Find yellow cells
    yellow_cells = find_yellow_cells(input_file)
    
    if len(yellow_cells) == 0:
        print("No yellow cells found!")
        return False
    
    # Step 2: Read the Excel file as data
    data = pd.read_excel(input_file)
    
    # Step 3: Find which columns contain which languages
    german_column = None
    french_column = None
    italian_column = None
    english_column = None
    
    # Look for German column
    for column_name in data.columns:
        if column_name.upper() == 'DE' or column_name.upper() == 'DEU':
            german_column = column_name
            break
    
    if german_column is None:
        print("Could not find German column (DE or DEU)")
        return False
    
    # Look for other language columns
    for column_name in data.columns:
        column_upper = column_name.upper()
        if column_upper == 'FR' or column_upper == 'FRA':
            french_column = column_name
        elif column_upper == 'IT' or column_upper == 'ITA':
            italian_column = column_name
        elif column_upper == 'EN' or column_upper == 'ENG':
            english_column = column_name
    
    # Step 4: Figure out column positions
    column_positions = {}
    position = 1
    for column_name in data.columns:
        column_positions[column_name] = position
        position = position + 1
    
    # Step 5: Process each row that has yellow cells
    results = []
    
    for row_num in yellow_cells:
        if row_num <= 1:  # Skip header row
            continue
            
        # Get the actual data row (Excel row 2 = data index 0)
        data_row_index = row_num - 2
        
        if data_row_index < 0 or data_row_index >= len(data):
            continue
            
        actual_row = data.iloc[data_row_index]
        
        # Create new row for output
        new_row = {}
        
        # Add Komponente if it exists
        if 'Komponente' in data.columns:
            komponente_value = actual_row['Komponente']
            if pd.notna(komponente_value) and str(komponente_value).strip() != '':
                new_row['Komponente'] = komponente_value
        
        # Always add German source
        new_row['Source'] = actual_row[german_column]
        
        # Add French if it was highlighted yellow
        if french_column:
            french_position = column_positions[french_column]
            if french_position in yellow_cells[row_num]:
                new_row['FR'] = yellow_cells[row_num][french_position]
            else:
                new_row['FR'] = ''
        
        # Add Italian if it was highlighted yellow
        if italian_column:
            italian_position = column_positions[italian_column]
            if italian_position in yellow_cells[row_num]:
                new_row['IT'] = yellow_cells[row_num][italian_position]
            else:
                new_row['IT'] = ''
        
        # Add English if it was highlighted yellow
        if english_column:
            english_position = column_positions[english_column]
            if english_position in yellow_cells[row_num]:
                new_row['EN'] = yellow_cells[row_num][english_position]
            else:
                new_row['EN'] = ''
        
        # Only add row if source text exists
        if new_row['Source'] and str(new_row['Source']).strip() != '':
            results.append(new_row)
    
    if len(results) == 0:
        print("No valid rows found")
        return False
    
    # Step 6: Save to new Excel file with formatting
    save_excel_with_formatting(results, input_file, output_file)
    
    print(f"Created {output_file} with {len(results)} rows")
    return True

def save_excel_with_formatting(results, input_file, output_file):
    """Save Excel file with text wrapping and column widths from original file"""
    
    # Create DataFrame
    final_data = pd.DataFrame(results)
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    final_data.to_excel(writer, index=False)
    
    # Get the worksheet we just created
    output_sheet = writer.sheets['Sheet1']
    
    # Open original file to copy column widths
    try:
        input_workbook = openpyxl.load_workbook(input_file)
        input_sheet = input_workbook.active
        
        # Copy column widths and set text wrapping
        column_number = 1
        for column_name in final_data.columns:
            # Get column letter (A, B, C, etc.)
            column_letter = openpyxl.utils.get_column_letter(column_number)
            
            # Copy column width from original file if it exists
            if column_letter in input_sheet.column_dimensions:
                original_width = input_sheet.column_dimensions[column_letter].width
                output_sheet.column_dimensions[column_letter].width = original_width
            else:
                # Set default width if original doesn't exist
                output_sheet.column_dimensions[column_letter].width = 20
            
            # Set text wrapping for all cells in this column
            for cell in output_sheet[column_letter]:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
            
            column_number = column_number + 1
        
        input_workbook.close()
        
    except Exception as e:
        print(f"Warning: Could not copy formatting: {e}")
        # Still set text wrapping even if width copying failed
        for column_number in range(1, len(final_data.columns) + 1):
            column_letter = openpyxl.utils.get_column_letter(column_number)
            output_sheet.column_dimensions[column_letter].width = 20
            for cell in output_sheet[column_letter]:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    
    # Save the file
    writer.close()

def get_output_filename(input_file):
    """Create output filename by adding _memoQ to input filename"""
    if input_file.endswith('.xlsx'):
        return input_file.replace('.xlsx', '_memoQ.xlsx')
    elif input_file.endswith('.xls'):
        return input_file.replace('.xls', '_memoQ.xlsx')
    else:
        return input_file + '_memoQ.xlsx'

def run_interactive_mode():
    """Run in interactive mode"""
    print("=== Excel Yellow Cell Extractor ===")
    print()
    
    # Get input file from user
    input_file = input("Enter the name of your Excel file (with .xlsx): ")
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found!")
        input("Press Enter to close...")
        return
    
    # Create output file name
    output_file = get_output_filename(input_file)
    
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print()
    
    # Do the work
    success = create_memoq_file(input_file, output_file)
    
    if success:
        print("✅ Done! Your file is ready for memoQ.")
    else:
        print("❌ Something went wrong. Please check your file.")
    
    input("Press Enter to close...")

def run_command_line_mode():
    """Run in command line mode"""
    if len(sys.argv) < 2:
        print("Usage: python excel_yellow_to_memoq.py <input_file> [output_file]")
        print("Example: python excel_yellow_to_memoq.py \"file.xlsx\"")
        print("Example: python excel_yellow_to_memoq.py \"file.xlsx\" \"output.xlsx\"")
        return
    
    input_file = sys.argv[1]
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found!")
        return
    
    # Get output file name
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        output_file = get_output_filename(input_file)
    
    print(f"Processing: {input_file}")
    print(f"Output: {output_file}")
    
    # Do the work
    success = create_memoq_file(input_file, output_file)
    
    if success:
        print(f"✅ SUCCESS! Processed file saved as {output_file}")
    else:
        print("❌ FAILED! Please check your file.")

# Main program - decide which mode to run
if len(sys.argv) > 1:
    # Command line mode - arguments were provided
    run_command_line_mode()
else:
    # Interactive mode - no arguments provided
    run_interactive_mode()
